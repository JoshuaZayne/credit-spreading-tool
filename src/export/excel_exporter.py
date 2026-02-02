"""
Excel Exporter Module
=====================

Exports credit analysis results to professionally formatted Excel workbooks.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exports credit analysis results to formatted Excel workbooks.

    Creates workbooks with multiple sheets:
        - Summary: Overview of all analyzed companies
        - UCA Cash Flow: Detailed cash flow analysis
        - DSCR Analysis: Debt service coverage with risk indicators
        - Leverage Ratios: Comprehensive leverage metrics
    """

    def __init__(self, risk_thresholds: Dict[str, Any] = None):
        """
        Initialize the Excel Exporter.

        Args:
            risk_thresholds: Dictionary containing risk threshold configurations
        """
        self.risk_thresholds = risk_thresholds or {}
        self._setup_styles()

    def _setup_styles(self):
        """Set up Excel cell styles."""
        # Colors
        self.colors = {
            "header_bg": "1F4E79",       # Dark blue
            "header_fg": "FFFFFF",       # White
            "healthy": "C6EFCE",         # Light green
            "warning": "FFEB9C",         # Light yellow
            "high_risk": "FFC7CE",       # Light red
            "neutral": "F2F2F2",         # Light gray
            "border": "B4B4B4"           # Gray border
        }

        # Border style
        self.thin_border = Border(
            left=Side(style='thin', color=self.colors["border"]),
            right=Side(style='thin', color=self.colors["border"]),
            top=Side(style='thin', color=self.colors["border"]),
            bottom=Side(style='thin', color=self.colors["border"])
        )

    def export(self, results: Dict[str, Any], output_path: str) -> str:
        """
        Export analysis results to Excel workbook.

        Args:
            results: Dictionary containing all analysis results
            output_path: Path for output Excel file

        Returns:
            Path to created Excel file
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, results)
        self._create_cash_flow_sheet(wb, results)
        self._create_dscr_sheet(wb, results)
        self._create_leverage_sheet(wb, results)

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        logger.info(f"Excel report saved to: {output_path}")
        return str(output_path)

    def _create_summary_sheet(self, wb: Workbook, results: Dict[str, Any]):
        """Create summary sheet with overview of all companies."""
        ws = wb.create_sheet("Summary")

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = "Credit Analysis Summary Report"
        ws["A1"].font = Font(bold=True, size=16, color=self.colors["header_fg"])
        ws["A1"].fill = PatternFill(
            start_color=self.colors["header_bg"],
            end_color=self.colors["header_bg"],
            fill_type="solid"
        )
        ws["A1"].alignment = Alignment(horizontal="center")

        # Report metadata
        ws["A3"] = "Report Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A4"] = "Total Companies:"
        ws["B4"] = results.get("summary", {}).get("total_companies", 0)

        # Summary statistics
        summary = results.get("summary", {})
        dscr_summary = summary.get("dscr", {})

        ws["A6"] = "DSCR Summary"
        ws["A6"].font = Font(bold=True)
        ws["A7"] = "Average DSCR:"
        ws["B7"] = f"{dscr_summary.get('average', 0):.2f}" if dscr_summary.get('average') else "N/A"
        ws["A8"] = "Min DSCR:"
        ws["B8"] = f"{dscr_summary.get('min', 0):.2f}" if dscr_summary.get('min') else "N/A"
        ws["A9"] = "Max DSCR:"
        ws["B9"] = f"{dscr_summary.get('max', 0):.2f}" if dscr_summary.get('max') else "N/A"
        ws["A10"] = "High Risk Count:"
        ws["B10"] = dscr_summary.get("high_risk_count", 0)

        # Company table headers
        headers = [
            "Ticker", "Revenue ($)", "EBITDA ($)", "Total Debt ($)",
            "DSCR", "DSCR Risk", "D/E Ratio", "D/EBITDA"
        ]

        header_row = 13
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors["header_fg"])
            cell.fill = PatternFill(
                start_color=self.colors["header_bg"],
                end_color=self.colors["header_bg"],
                fill_type="solid"
            )
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center")

        # Company data
        row = header_row + 1
        for company in results.get("companies", []):
            financial = company.get("financial_data", {})
            ticker = company.get("ticker", "")

            # Find DSCR for this company
            dscr_data = next(
                (d for d in results.get("dscr_results", []) if d.get("ticker") == ticker),
                {}
            )
            dscr_result = dscr_data.get("dscr", {})

            # Find leverage ratios
            leverage_data = next(
                (l for l in results.get("leverage_ratios", []) if l.get("ticker") == ticker),
                {}
            )
            ratios = leverage_data.get("ratios", {})

            # Write data
            ws.cell(row=row, column=1, value=ticker).border = self.thin_border
            ws.cell(row=row, column=2, value=financial.get("revenue", 0)).border = self.thin_border
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3, value=financial.get("ebitda", 0)).border = self.thin_border
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4, value=financial.get("total_debt", 0)).border = self.thin_border
            ws.cell(row=row, column=4).number_format = '#,##0'

            # DSCR
            dscr_cell = ws.cell(row=row, column=5, value=dscr_result.get("value"))
            dscr_cell.number_format = '0.00'
            dscr_cell.border = self.thin_border

            # DSCR Risk with conditional formatting
            risk_cell = ws.cell(row=row, column=6, value=dscr_result.get("risk_level", "N/A"))
            risk_cell.border = self.thin_border
            self._apply_risk_formatting(risk_cell, dscr_result.get("risk_level"))

            # D/E Ratio
            de_ratio = ratios.get("debt_to_equity", {}).get("value")
            ws.cell(row=row, column=7, value=de_ratio).border = self.thin_border
            ws.cell(row=row, column=7).number_format = '0.00'

            # D/EBITDA
            d_ebitda = ratios.get("debt_to_ebitda", {}).get("value")
            ws.cell(row=row, column=8, value=d_ebitda).border = self.thin_border
            ws.cell(row=row, column=8).number_format = '0.00'

            row += 1

        # Adjust column widths
        self._auto_adjust_columns(ws)

    def _create_cash_flow_sheet(self, wb: Workbook, results: Dict[str, Any]):
        """Create UCA Cash Flow analysis sheet."""
        ws = wb.create_sheet("UCA Cash Flow")

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "UCA Cash Flow Analysis"
        ws["A1"].font = Font(bold=True, size=14, color=self.colors["header_fg"])
        ws["A1"].fill = PatternFill(
            start_color=self.colors["header_bg"],
            end_color=self.colors["header_bg"],
            fill_type="solid"
        )

        # Headers
        headers = [
            "Ticker", "Operating Cash Flow", "Free Cash Flow",
            "Debt Service Req.", "Cash for Debt", "NOI", "Coverage Ratio"
        ]

        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors["header_fg"])
            cell.fill = PatternFill(
                start_color=self.colors["header_bg"],
                end_color=self.colors["header_bg"],
                fill_type="solid"
            )
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data
        row = header_row + 1
        for cf_data in results.get("cash_flows", []):
            ticker = cf_data.get("ticker", "")
            cash_flow = cf_data.get("cash_flow", {})

            ws.cell(row=row, column=1, value=ticker).border = self.thin_border

            for col, key in enumerate([
                "operating_cash_flow", "free_cash_flow", "debt_service_requirement",
                "cash_available_for_debt", "noi", "uca_coverage_ratio"
            ], 2):
                cell = ws.cell(row=row, column=col, value=cash_flow.get(key))
                cell.border = self.thin_border
                if key == "uca_coverage_ratio":
                    cell.number_format = '0.00'
                else:
                    cell.number_format = '#,##0'

            row += 1

        self._auto_adjust_columns(ws)

    def _create_dscr_sheet(self, wb: Workbook, results: Dict[str, Any]):
        """Create DSCR analysis sheet."""
        ws = wb.create_sheet("DSCR Analysis")

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = "Debt Service Coverage Ratio (DSCR) Analysis"
        ws["A1"].font = Font(bold=True, size=14, color=self.colors["header_fg"])
        ws["A1"].fill = PatternFill(
            start_color=self.colors["header_bg"],
            end_color=self.colors["header_bg"],
            fill_type="solid"
        )

        # Headers
        headers = [
            "Ticker", "DSCR", "Risk Level", "NOI ($)",
            "Interest ($)", "Principal ($)", "Total Debt Service ($)", "Interpretation"
        ]

        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors["header_fg"])
            cell.fill = PatternFill(
                start_color=self.colors["header_bg"],
                end_color=self.colors["header_bg"],
                fill_type="solid"
            )
            cell.border = self.thin_border

        # Data
        row = header_row + 1
        for dscr_data in results.get("dscr_results", []):
            ticker = dscr_data.get("ticker", "")
            dscr = dscr_data.get("dscr", {})
            details = dscr.get("details", {})

            ws.cell(row=row, column=1, value=ticker).border = self.thin_border

            dscr_cell = ws.cell(row=row, column=2, value=dscr.get("value"))
            dscr_cell.number_format = '0.00'
            dscr_cell.border = self.thin_border

            risk_cell = ws.cell(row=row, column=3, value=dscr.get("risk_level", "N/A"))
            risk_cell.border = self.thin_border
            self._apply_risk_formatting(risk_cell, dscr.get("risk_level"))

            for col, key in enumerate([
                "net_operating_income", "interest_expense",
                "principal_payment", "total_debt_service"
            ], 4):
                cell = ws.cell(row=row, column=col, value=details.get(key))
                cell.number_format = '#,##0'
                cell.border = self.thin_border

            interp_cell = ws.cell(row=row, column=8, value=dscr.get("interpretation", ""))
            interp_cell.border = self.thin_border
            interp_cell.alignment = Alignment(wrap_text=True)

            row += 1

        self._auto_adjust_columns(ws)
        ws.column_dimensions['H'].width = 50  # Wider column for interpretation

    def _create_leverage_sheet(self, wb: Workbook, results: Dict[str, Any]):
        """Create leverage ratios sheet."""
        ws = wb.create_sheet("Leverage Ratios")

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = "Leverage Ratio Analysis"
        ws["A1"].font = Font(bold=True, size=14, color=self.colors["header_fg"])
        ws["A1"].fill = PatternFill(
            start_color=self.colors["header_bg"],
            end_color=self.colors["header_bg"],
            fill_type="solid"
        )

        # Headers
        headers = [
            "Ticker", "D/E Ratio", "D/E Risk", "D/EBITDA", "D/EBITDA Risk",
            "Int. Coverage", "Int. Cov. Risk", "Total Debt ($)", "Total Equity ($)"
        ]

        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color=self.colors["header_fg"])
            cell.fill = PatternFill(
                start_color=self.colors["header_bg"],
                end_color=self.colors["header_bg"],
                fill_type="solid"
            )
            cell.border = self.thin_border

        # Data
        row = header_row + 1
        for leverage_data in results.get("leverage_ratios", []):
            ticker = leverage_data.get("ticker", "")
            ratios = leverage_data.get("ratios", {})

            # Find financial data for this company
            company = next(
                (c for c in results.get("companies", []) if c.get("ticker") == ticker),
                {}
            )
            financial = company.get("financial_data", {})

            ws.cell(row=row, column=1, value=ticker).border = self.thin_border

            # D/E Ratio
            de = ratios.get("debt_to_equity", {})
            ws.cell(row=row, column=2, value=de.get("value")).border = self.thin_border
            ws.cell(row=row, column=2).number_format = '0.00'
            de_risk_cell = ws.cell(row=row, column=3, value=de.get("risk_level"))
            de_risk_cell.border = self.thin_border
            self._apply_risk_formatting(de_risk_cell, de.get("risk_level"))

            # D/EBITDA
            d_ebitda = ratios.get("debt_to_ebitda", {})
            ws.cell(row=row, column=4, value=d_ebitda.get("value")).border = self.thin_border
            ws.cell(row=row, column=4).number_format = '0.00'
            d_ebitda_risk_cell = ws.cell(row=row, column=5, value=d_ebitda.get("risk_level"))
            d_ebitda_risk_cell.border = self.thin_border
            self._apply_risk_formatting(d_ebitda_risk_cell, d_ebitda.get("risk_level"))

            # Interest Coverage
            int_cov = ratios.get("interest_coverage", {})
            ws.cell(row=row, column=6, value=int_cov.get("value")).border = self.thin_border
            ws.cell(row=row, column=6).number_format = '0.00'
            int_cov_risk_cell = ws.cell(row=row, column=7, value=int_cov.get("risk_level"))
            int_cov_risk_cell.border = self.thin_border
            self._apply_risk_formatting(int_cov_risk_cell, int_cov.get("risk_level"))

            # Financial data
            ws.cell(row=row, column=8, value=financial.get("total_debt")).border = self.thin_border
            ws.cell(row=row, column=8).number_format = '#,##0'
            ws.cell(row=row, column=9, value=financial.get("total_equity")).border = self.thin_border
            ws.cell(row=row, column=9).number_format = '#,##0'

            row += 1

        self._auto_adjust_columns(ws)

    def _apply_risk_formatting(self, cell, risk_level: str):
        """Apply color formatting based on risk level."""
        color_map = {
            "LOW": self.colors["healthy"],
            "MEDIUM": self.colors["warning"],
            "HIGH": self.colors["high_risk"],
            "CRITICAL": self.colors["high_risk"],
            "N/A": self.colors["neutral"]
        }

        color = color_map.get(risk_level, self.colors["neutral"])
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
